import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl import Workbook
import pytz


def format_excel_data(feedback_and_rate_result, data_excel_path, file_prefix):
    # Формируем DataFrame из данных
    data = {
        'Артикул': [art['art'] for art in feedback_and_rate_result],
        'Рейтинг': [],
        'Количество отзывов': [],
    }

    # Пытаемся получить значения по ключу 'rating_art' и 'feedbacks_count'
    for art in feedback_and_rate_result:
        rating_art = art.get('rating_art')
        if rating_art is not None:
            data['Рейтинг'].append(rating_art)
        else:
            data['Рейтинг'].append('')

        feedbacks_count = art.get('feedbacks_count')
        if feedbacks_count is not None:
            data['Количество отзывов'].append(feedbacks_count)
        else:
            data['Количество отзывов'].append(
                '')

    # Создадим список списков для оценок последних отзывов
    rate_feedbacks_list = [art.get('rate_feedbacks', []) for art in feedback_and_rate_result]
    # Заполним список списков оценками последних отзывов, добавляя None, если оценок меньше 5
    rate_feedbacks_list = [rate + [None] * (5 - len(rate)) for rate in rate_feedbacks_list]
    # Развернем список списков, чтобы оценки были в отдельных столбцах
    rate_feedbacks_list_transposed = list(map(list, zip(*rate_feedbacks_list)))

    # Создаем столбцы с оценками
    for i in range(1, 6):
        data[f'Оценка {i}'] = rate_feedbacks_list_transposed[i - 1]

    df = pd.DataFrame(data)

    # Создаем или открываем существующий Excel-файл
    wb = Workbook()
    ws = wb.active

    # Заполняем лист данными из DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Список цветов для оценок
    rating_colors = {
        1: "ffbe2f33",
        2: "ffff6840",
        3: "ffffb173",
        4: "ffcaf86f",
        5: "ff689c00"
    }

    # Проходимся по столбцам с оценками и устанавливаем цвет ячеек
    for col_num in range(4, 9):  # Начиная со столбца "D" и до столбца "H"
        col_letter = openpyxl.utils.get_column_letter(col_num)

        # Проходимся по каждой ячейке в столбце
        for index, cell in enumerate(ws[col_letter]):
            if index > 0:  # Пропускаем первую ячейку с заголовком
                rating = cell.value
                if rating in rating_colors:  # Проверяем, есть ли оценка в списке цветов
                    cell.fill = PatternFill(start_color=rating_colors[rating], end_color=rating_colors[rating],
                                            fill_type="solid")

    # Получаем все ячейки в листе
    all_cells = ws['A1':openpyxl.utils.get_column_letter(len(df.columns)) + str(len(df) + 1)]

    # Устанавливаем шрифт Arial и размер 10 для всех ячеек
    font = Font(name='Arial', size=10)
    for row in all_cells:
        for cell in row:
            cell.font = font

    # Формируем имя файла в соответствии с схемой '{префикс}_{дата_создания_файла}.xlsx'
    tz = pytz.timezone('Europe/Moscow')
    current_datetime = datetime.now(tz)
    date_format = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"{file_prefix}_{date_format}.xlsx"
    excel_file_path = os.path.join(data_excel_path, file_name)

    # Сохраняем Excel-файл
    wb.save(excel_file_path)

    # Закрываем Workbook
    wb.close()

    # Отправляем файл пользователю
    return excel_file_path
